import os
import json
import logging
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

def merge_metadata_files(destination_folder):
    root_metadata_file = os.path.join(destination_folder, "metadata.json")
    
    try:
        merged_data = {}
        files_merged = False
        files_to_delete = []
        
        for root, dirs, files in os.walk(destination_folder):
            if "metadata.json" in files:
                metadata_file_path = os.path.join(root, "metadata.json")
                relative_path = os.path.relpath(root, destination_folder)
                depth = len(relative_path.split(os.sep))
                
                try:
                    with open(metadata_file_path, 'r', encoding='utf-8') as sub_file:
                        file_content = sub_file.read()
                        cleaned_content = ''.join(char for char in file_content if ord(char) >= 32 or char in '\n\r\t')
                        sub_data = json.loads(cleaned_content)
                        if sub_data:
                            # Merge data into the appropriate level
                            current_level = merged_data
                            path_parts = relative_path.split(os.sep)
                            for part in path_parts[:-1]:
                                if part not in current_level:
                                    current_level[part] = {}
                                current_level = current_level[part]
                            current_level.update(sub_data)
                            files_merged = True
                    
                    # Mark files in deeper levels for deletion
                    if depth > 1:
                        files_to_delete.append(metadata_file_path)
                
                except json.JSONDecodeError as json_err:
                    logging.warning(f"JSON error in {metadata_file_path}: {str(json_err)}. Skipping this file.")
                except Exception as e:
                    logging.error(f"Error processing {metadata_file_path}: {str(e)}")
        
        if files_merged:
            with open(root_metadata_file, 'w', encoding='utf-8') as root_file:
                json.dump(merged_data, root_file, indent=2)
        
        # Delete metadata files in deeper levels
        for file_path in files_to_delete:
            try:
                os.remove(file_path)
            except Exception as e:
                logging.error(f"Error deleting {file_path}: {str(e)}")
        
    except Exception as e:
        logging.error(f"Error merging and deleting metadata files: {str(e)}")

def create_metadata_csv(destination_folder):
    try:
        metadata_file_path = os.path.join(destination_folder, "metadata.json")
        merged_metadata = {}
    
        for root, dirs, files in os.walk(destination_folder):
            if "metadata.json" in files:
                sub_metadata_path = os.path.join(root, "metadata.json")
                try:
                    with open(sub_metadata_path, 'r', encoding='utf-8') as f:
                        sub_metadata = json.load(f)
                    relative_path = os.path.relpath(root, destination_folder)
                    merged_metadata[relative_path] = sub_metadata
                except json.JSONDecodeError:
                    logging.warning(f"Invalid JSON in {sub_metadata_path}. Skipping this file.")
                except Exception as e:
                    logging.error(f"Error reading {sub_metadata_path}: {str(e)}")


        if os.path.exists(metadata_file_path) and os.path.getsize(metadata_file_path) > 0:
            try:
                with open(metadata_file_path, 'r', encoding='utf-8') as f:
                    main_metadata = json.load(f)
                merged_metadata.update(main_metadata)
            except json.JSONDecodeError:
                logging.warning(f"Invalid JSON in main metadata file {metadata_file_path}. Skipping this file.")
            except Exception as e:
                logging.error(f"Error reading main metadata file {metadata_file_path}: {str(e)}")

        if not merged_metadata:
            logging.warning("No metadata found. Creating an empty Excel file.")

        # Change output to Excel
        output_path = os.path.join(destination_folder, "all_metadata.xlsx")
        workbook = Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        def humanize_key(key):
            # Remove any prefix dots
            key = key.lstrip('.')
            # Split on dots, underscores, and camelCase
            words = []
            current_word = ''
            
            for char in key:
                if char in '._':
                    if current_word:
                        words.append(current_word)
                    current_word = ''
                elif char.isupper() and current_word and current_word[-1].islower():
                    words.append(current_word)
                    current_word = char
                else:
                    current_word += char
            
            if current_word:
                words.append(current_word)
            
            # Capitalize first letter of each word and join with spaces
            return ' '.join(word.capitalize() for word in words)

        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        regular_font = Font(name='Arial', size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        for folder_path, files in merged_metadata.items():
            for filename, file_metadata in files.items():
                if isinstance(file_metadata, str):
                    logging.warning(f"Unexpected string value for file_metadata: {file_metadata}")
                    continue
                
                # Create a new sheet for each file
                safe_sheet_name = filename[:31]
                worksheet = workbook.create_sheet(title=safe_sheet_name)

                # Flatten the metadata
                def flatten_metadata(metadata, prefix=''):
                    if not isinstance(metadata, dict):
                        return {prefix.rstrip('.'): metadata}
                    flattened = {}
                    for key, value in metadata.items():
                        if isinstance(value, dict):
                            flattened.update(flatten_metadata(value, f"{prefix}{key}."))
                        elif not isinstance(value, list):
                            flattened[key] = value
                    return flattened
                
                # Add title headers
                worksheet.cell(row=1, column=1, value='Property')
                worksheet.cell(row=1, column=2, value='Value')
                
                # Style headers
                for col in [1, 2]:
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = header_alignment
                
                # Set row height for header
                worksheet.row_dimensions[1].height = 20
                
                flattened_metadata = flatten_metadata(file_metadata)
                
                # Write and style data rows
                row = 2  # Start after header
                for key, value in flattened_metadata.items():
                    # Write data
                    key_cell = worksheet.cell(row=row, column=1, value=humanize_key(key))
                    value_cell = worksheet.cell(row=row, column=2, value=str(value))
                    
                    # Style data cells
                    for cell in [key_cell, value_cell]:
                        cell.font = regular_font
                        cell.border = thin_border
                        cell.alignment = cell_alignment
                    
                    # Alternate row colors
                    if row % 2 == 0:
                        for col in [1, 2]:
                            worksheet.cell(row=row, column=col).fill = PatternFill(
                                start_color='F2F2F2',
                                end_color='F2F2F2',
                                fill_type='solid'
                            )
                    
                    row += 1
                
                # Adjust column widths
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column  # Get the column name
                    
                    # Find the maximum length of the cell value
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    
                    # Set width with some padding (min 10, max 50 characters)
                    adjusted_width = min(max(max_length + 2, 10), 50)
                    worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width
                
                # Freeze the header row
                worksheet.freeze_panes = 'A2'
                
                # Set zoom level
                worksheet.sheet_view.zoomScale = 100

        # Remove empty sheet if it exists
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])

        workbook.save(output_path)

        # Clean up root metadata file
        root_metadata_file = os.path.join(destination_folder, "metadata.json")
        if os.path.exists(root_metadata_file):
            try:
                os.remove(root_metadata_file)
            except Exception as e:
                logging.error(f"Error deleting root metadata file {root_metadata_file}: {str(e)}")
                
    except Exception as e:
        logging.error(f"Error creating Excel file from metadata: {str(e)}")
        logging.exception("Detailed traceback:")